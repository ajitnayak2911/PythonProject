import asyncio
import json
import random
import string
import openpyxl
import os
import glob
from datetime import datetime
from playwright.async_api import async_playwright
from urllib.parse import urlparse, parse_qs
from urllib.parse import urlsplit, urlunsplit


# =====================
# CONFIG
# =====================
DEV_USERNAME = "broadridgedigital"     # <-- put your dev username
DEV_PASSWORD = "broadridge1" # <-- put your dev password

INPUT_FILE = r"C:\Users\ajitk\Python_files\input.xlsx"
OUTPUT_FILE = r"C:\Users\ajitk\Python_files\output.xlsx"

# =====================
# HELPERS
# =====================

PARAM_COLS = [
    "utm_medium", "utm_source", "utm_campaign",
    "utm_term", "utm_content", "content_id",
    "campaign_id", "sub_source"
]

def apply_dev_auth(url: str) -> str:
    """Inject basic auth if URL is dev environment"""
    if "dev" in url or "www-dev" in url:
        parts = url.split("://")
        if len(parts) == 2 and "@" not in parts[1]:
            return f"{parts[0]}://{DEV_USERNAME}:{DEV_PASSWORD}@{parts[1]}"
    return url


def generate_dynamic_value(field_name: str, counter: int) -> str:
    """Generate dynamic values; Last Name always fixed TESTTEST."""
    name = field_name.lower()
    if "last" in name:
        return "TESTTEST"
    elif "first" in name:
        return f"FirstNameTest{counter}"
    elif "email" in name:
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        return f"autotest{counter}_{timestamp}@example.com"
    elif "phone" in name:
        return "".join([str(random.randint(0, 9)) for _ in range(10)])
    elif "company" in name:
        return f"TestCompany{counter}"
    elif "job" in name:
        return f"QA Engineer {counter}"
    elif "comment" in name or "message" in name:
        return "Automated test inquiry message with enough length (>25 chars)."
    else:
        return "".join(random.choices(string.ascii_letters, k=8))


def compare_payload(expected: dict, submitted: dict):
    """Compare expected vs submitted payload, tolerate aliases."""
    if not isinstance(submitted, dict):
        try:
            submitted = json.loads(submitted)
        except Exception:
            submitted = {}

    exp_norm = {k.lower(): str(v).strip().lower() for k, v in expected.items()}
    subm_norm = {k.lower(): str(v).strip().lower() for k, v in submitted.items()}

    aliases = {
        "company": ["company", "Company"],
        "phone_business": ["phone", "phone_business", "telephone"],
        "email_work": ["email", "email_work"],
        "name_first": ["first_name", "name_first"],
        "name_last": ["last_name", "name_last"],
        "comment": ["comment", "message"]
    }

    mismatches, matched = [], 0
    for k, v in exp_norm.items():
        if k in subm_norm and subm_norm[k] == v:
            matched += 1
            continue
        if k in aliases:
            if any(alias.lower() in subm_norm and subm_norm[alias.lower()] == v for alias in aliases[k]):
                matched += 1
                continue
        mismatches.append(f"{k} mismatch or missing")

    if matched >= max(1, len(exp_norm)//2):
        return "PASS", "All fields matched" if not mismatches else "; ".join(mismatches)
    return "FAIL", "; ".join(mismatches)

def validate_url_params_in_json(url: str, json_dict: dict):
        """
        Extract URL query parameters and check if all are present
        in the JSON response with matching values.
        Returns (overall_status, dict_of_param_results)
        """
        # protect against non‑dict inputs
        if not isinstance(json_dict, dict):
            try:
                json_dict = json.loads(json_dict)
            except Exception:
                json_dict = {}

        # parse ?key=value&... portion of URL
        parsed = urlparse(url)
        qparams = parse_qs(parsed.query)  # -> {"utm_medium": ["AffiliateMarketing"], ...}

        results = {}
        for key, vals in qparams.items():
            url_val = vals[0] if vals else ""
            match = False
            # compare ignoring case/whitespace
            for jk, jv in json_dict.items():
                if jk.lower() == key.lower() and str(jv).strip().lower() == url_val.strip().lower():
                    match = True
                    break
            results[key] = "PASS" if match else "FAIL"

        overall = "PASS" if all(v == "PASS" for v in results.values()) else "FAIL"
        return overall, results

async def process_form_submission(page, url: str, counter: int):
    url = apply_dev_auth(url)
    payloads = {}
    # valid initial dictionary; no ellipses
    extra_values = {
        "formSubmissionId": None,
        "fullURL": None,
        "page_id": None,
        "raw_response": None,
        "form_data_subset": {}
    }
    # --- capture requests (for payload comparison) ---
    # --- capture requests (for payload comparison) ---
    def capture_request(req):
        if req.method != "POST":
            return
        try:
            data = req.post_data_json
            if data:
                payloads["last"] = data
                return
        except Exception:
            pass
        try:
            raw = req.post_data
            if raw and "=" in raw:
                d = {}
                for pair in raw.split("&"):
                    k, _, v = pair.partition("=")
                    d[k] = v
                payloads["last"] = d
            else:
                payloads["last"] = raw or {}
        except:
            payloads["last"] = {}

    page.on("request", capture_request)

    # ✅ Capture form-processor response
    async def handle_response(res):
        try:
            if res.request.method == "POST" and "form-processor" in res.url.lower():
                print(f"🔎 Captured Response URL: {res.url}")
                raw_text = await res.text()
                extra_values["raw_response"] = {"raw_text": raw_text}

                # --- 🆕 Extract the 8 required parameters from "Form Data" ---
                interest_fields = [
                    "utm_medium", "utm_source", "utm_campaign",
                    "utm_term", "utm_content", "content_id",
                    "campaign_id", "sub_source"
                ]
                field_values = {}

                if "Form Data:" in raw_text:
                    try:
                        form_section = raw_text.split("Form Data:", 1)[1]
                        pairs = form_section.strip().split("&")
                        field_values = {}
                        for p in pairs:
                            k, _, v = p.partition("=")
                            if not k:
                                continue
                            key = k.strip()
                            val = v.strip()

                            # normalize obvious variants
                            norm = (
                                key.lower()
                                .replace("persistent_", "")
                                .replace("session_", "")
                                .replace("sub-source", "sub_source")
                            )
                            if norm in PARAM_COLS:
                                # 🆕  only take the first non‑empty value
                                if norm not in field_values or not field_values[norm]:
                                    field_values[norm] = val

                        extra_values["form_data_subset"] = field_values

                    except Exception as e:
                        print("⚠ Error parsing Form Data block:", e)
                # keep your existing JSON capture for diagnostics
                if "Request Body:" in raw_text:
                    try:
                        start = raw_text.index("Request Body:") + len("Request Body:")
                        trimmed = raw_text[start:].strip()
                        if trimmed.startswith("{"):
                            json_str = trimmed.split("}", 1)[0] + "}"
                            data = json.loads(json_str)
                            extra_values["raw_parsed"] = data
                            extra_values["formSubmissionId"] = data.get("formSubmissionId", "")
                            extra_values["fullURL"] = data.get("fullURL", "")
                            extra_values["page_id"] = data.get("page_id", "")
                    except Exception as e:
                        print("⚠ Error extracting Request Body JSON:", e)
        except Exception as e:
            print("⚠ Error in handle_response:", e)
    page.on("response", lambda res: asyncio.create_task(handle_response(res)))

    await page.goto(url, timeout=60000)
    # Cookie banner
    try:
        await page.click("#onetrust-accept-btn-handler", timeout=5000)
    except:
        try:
            await page.click("#onetrust-close-btn-container button", timeout=5000)
        except:
            pass

    await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    await page.wait_for_timeout(2000)

    # --- form detection with fallback ---
    form, form_source = None, "none"
    if await page.locator("form.contact-us__form[data-tracker-identifier='Page bottom form']").count() > 0:
        form = page.locator("form.contact-us__form[data-tracker-identifier='Page bottom form']").first
        form_source = "bottom"
    elif await page.locator("div.nav-cta >> button.modal-trigger").count() > 0:
        await page.click("div.nav-cta >> button.modal-trigger")
        try:
            await page.wait_for_selector("form.contact-us__form", state="visible", timeout=20000)
            form = page.locator("form.contact-us__form").first
            form_source = "modal"
        except:
            pass

    if not form:
        return "ERROR", {}, {}, "No form found", "No Thank You", form_source, None, None, None, None

    # --- fill fields ---
    filled_data = {}
    for fname in ["name_first", "name_last", "email_work", "phone_business", "job_title", "Company"]:
        val = generate_dynamic_value(fname, counter)
        await form.locator(f"input[name='{fname}']").fill(val)
        filled_data[fname] = val

    # Country dropdown
    try:
        await form.locator("button.dropdown-trigger").click()
        options = page.locator("li.dropdown-item:visible")
        if await options.count() > 0:
            idx = random.randint(0, await options.count() - 1)
            option = options.nth(idx)
            option_val = await option.get_attribute("data-value")
            await option.click()
            filled_data["country"] = option_val or "Unknown"
    except:
        filled_data["country"] = "Not selected"

    # Comment
    comment_val = generate_dynamic_value("comment", counter)
    await form.locator("textarea[name='comment']").fill(comment_val)
    filled_data["comment"] = comment_val

    # DOM fallback
    form_values = await form.evaluate("""(form) => {
        const d = {};
        form.querySelectorAll("input, textarea, select").forEach(el => {
            if (el.name && el.type !== "hidden") {
                d[el.name] = el.value || "";
            }
        });
        return d;
    }""")

    # --- submit ---
    confirmation_text = "No Thank You message found"
    try:
        await form.locator("button.contact-us__form-button[type='submit']").click()
        await page.wait_for_timeout(8000)  # wait for network responses

        # Confirmation text
        try:
            success_locator = page.locator("div.contact-us__success")
            await success_locator.wait_for(state="visible", timeout=8000)
            confirmation_text = await success_locator.inner_text()
        except:
            pass

    except Exception as e:
        return "ERROR", filled_data, {}, f"No submit button: {e}", confirmation_text, form_source, None, None, None, None

    # --- submission data ---
    submitted = payloads.get("last", {}) or {}
    if not isinstance(submitted, dict):
        submitted = {}

    important_keys = ["name_first", "name_last", "email", "phone", "company", "comment"]
    if not any(any(key in k.lower() for key in important_keys) for k in submitted.keys()):
        submitted = form_values

    # Compare what we filled vs what we submitted
    result, notes = compare_payload(filled_data, submitted)

    # ✅ NEW CODE — capture the 8 parameter values extracted from Form Data
    param_results = {}
    form_subset = extra_values.get("form_data_subset", {})

    # Only keep the 8 required fields
    for key in PARAM_COLS:
        # get() returns "" if missing
        param_results[key] = form_subset.get(key, "")

    # Add the captured values to notes (for reference in Excel)
    notes = f"{notes}; FormParameters={json.dumps(param_results, ensure_ascii=False)}"

    # --- Return collected results ---
    return (
        result,
        filled_data,
        submitted,
        notes,
        confirmation_text,
        form_source,
        extra_values.get("formSubmissionId"),
        extra_values.get("fullURL"),
        extra_values.get("page_id"),
        extra_values
        #extra_values.get("raw_response"),
    )
# =====================
# MAIN DRIVER
# =====================

async def main():
    wb = openpyxl.load_workbook(INPUT_FILE)
    sheet = wb.active

    # add the parameter‑level columns
    PARAM_COLS = [
        "utm_medium", "utm_source", "utm_campaign",
        "utm_term", "utm_content", "content_id",
        "campaign_id", "sub_source"
    ]

    headers = [
                  "URL", "Result", "Filled Fields", "Captured Payload", "Notes",
                  "Confirmation", "Form Source", "FormSubmissionId",
                  "fullURL", "page_id", "Raw JSON Response"
              ] + PARAM_COLS + ["Overall Result"]
    for idx, name in enumerate(headers, start=1):
        if sheet.cell(row=1, column=idx).value != name:
            sheet.cell(row=1, column=idx, value=name)

    url_col = headers.index("URL") + 1
    result_col = headers.index("Result") + 1
    filled_col = headers.index("Filled Fields") + 1
    payload_col = headers.index("Captured Payload") + 1
    notes_col = headers.index("Notes") + 1
    confirm_col = headers.index("Confirmation") + 1
    source_col = headers.index("Form Source") + 1
    fsid_col = headers.index("FormSubmissionId") + 1
    fullurl_col = headers.index("fullURL") + 1
    pageid_col = headers.index("page_id") + 1
    rawjson_col = headers.index("Raw JSON Response") + 1

    async with async_playwright() as p:
        # ✅ Auto-detect latest Chromium binary
        browser = await p.chromium.launch(headless=True)

        # ✅ NEW CODE — use a single browser context for all forms
        context = await browser.new_context()
        page = await context.new_page()

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            url = row[url_col - 1]
            print(f"Row {i} -> {url}")
            current_result = row[result_col - 1]
            if not url:
                continue
            #if current_result and str(current_result).strip():
                #print(f"⚡ Skipping row {i} (already has result: {current_result})")
                #continue

            print(f"▶ Testing: {url}")
            context = await browser.new_context()
            page = await context.new_page()
            try:
                result, filled_data, submitted, notes, confirm, form_source, form_submission_id, \
                    full_url_val, page_id_val, extra_data = await process_form_submission(page, url, i)
                sheet.cell(row=i, column=result_col).value = result
                sheet.cell(row=i, column=filled_col).value = json.dumps(filled_data)
                sheet.cell(row=i, column=payload_col).value = json.dumps(submitted)
                sheet.cell(row=i, column=notes_col).value = notes
                sheet.cell(row=i, column=confirm_col).value = confirm
                sheet.cell(row=i, column=source_col).value = form_source
                sheet.cell(row=i, column=fsid_col).value = form_submission_id
                sheet.cell(row=i, column=fullurl_col).value = full_url_val
                sheet.cell(row=i, column=pageid_col).value = page_id_val
                raw_response = extra_data.get("raw_response")
                sheet.cell(row=i, column=rawjson_col).value = json.dumps(raw_response) if raw_response else None

                # --- Write parameter columns ---
                param_dict = {}
                try:
                    if isinstance(extra_data, dict) and "form_data_subset" in extra_data:
                        param_dict = extra_data["form_data_subset"]
                    elif "FormParameters=" in notes:
                        form_params_json = notes.split("FormParameters=", 1)[-1]
                        param_dict = json.loads(form_params_json)
                except Exception as e:
                    print(f"⚠ Error reading parameter dictionary on row {i}: {e}")
                    param_dict={}

                #start_col = len(headers) - len(PARAM_COLS) + 1
                utmmedium_col = headers.index("utm_medium") + 1

                for col_offset, param_name in enumerate(PARAM_COLS):
                    sheet.cell(row=i, column=utmmedium_col + col_offset).value = param_dict.get(param_name, "")

                # --- Compute Overall Result ---
                try:
                    url_value = str(url).strip()
                    fullurl_value = str(full_url_val).strip()
                    overall_col = len(headers)

                    def normalize(u: str):
                        parts = list(urlsplit(u.strip()))
                        if parts[2].endswith("/"):
                            parts[2] = parts[2].rstrip("/")
                        return urlunsplit(parts).lower()

                    if normalize(url_value) == normalize(fullurl_value):
                        sheet.cell(row=i, column=overall_col).value = "PASS"
                    else:
                        sheet.cell(row=i, column=overall_col).value = "FAIL"
                except Exception as e:
                    print(f"⚠ Error computing Overall Result on row {i}: {e}")
                    sheet.cell(row=i, column=len(headers)).value = "ERROR"
            except Exception as e:
                sheet.cell(row=i, column=result_col).value = "ERROR"
                sheet.cell(row=i, column=notes_col).value = str(e)
            finally:
                await context.close()



        await browser.close()

    wb.save(OUTPUT_FILE)
    print(f"✅ Results saved in {OUTPUT_FILE}")


if __name__ == "__main__":
    asyncio.run(main())